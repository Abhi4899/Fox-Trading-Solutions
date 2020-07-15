#!/usr/bin/env python
# coding: utf-8

# In[1]:


import sys
import random
import openpyxl
import urllib.request, urllib.parse, urllib.error
import requests
import json


# In[2]:


#opening the spreadsheet to work with
wb = openpyxl.load_workbook('spreadsheet.xlsx')


# In[3]:


#starting to work on sheet1
sheet=wb['Sheet1']


# In[4]:


#to create the heading of the spreadsheet
headings = ['City','Temperature','Humidity','Unit','Upload (0/1)']
for i in range(len(headings)):
    sheet.cell(row=1, column=i+1).value=headings[i]


# In[5]:


#saving the changes
wb.save('task1.xlsx')


# In[6]:


print("""Enter city names with you preferred temperature and
do you want to refresh per sec as shown below:
<CityName><space><TemperatureFormat><space><Upadate/NoUpdate>
Delhi C 1
Kolkatta F 0
here, 0 - updates per second
      1 - does not updates per second
      F - temperature in farenhite
      C - temperature in degree celsius
Press enter to stop""")

#variable named data is a dictionary with city as key and a list as value
#the list stores [<TemperatureFormat>,<Upadate/NoUpdate>]
data = {}
while True:
    s=input()
    
    #if nothing is entered, the loop will terminate
    if len(s)<1:
        break
    
    l=s.split() #list of strings
    
    #if the length of the sring is less than required insted of stopping the execution
    #if will ask to input in the correct format
    if len(l)<3:
        print('Please enter in valid format')
        
    update=int(l[-1]) #stores the update value
    
    unit=l[-2] #stores the unit C/F
    
    l=l[:-2] #removing update and unit values so that remaing can be joined to form the city name
    
    city=' '.join(l) #creating city name from remaining elements of the list 'l'
    
    #checking the conditions as sated in the print statement of input instructions
    if update not in [0,1]:
        print('Please enter in valid format')
    elif unit not in ['C','F','c','f']:
        print('Please enter in valid format')
    else:
        data[city]=[unit.upper(),int(update)]


# In[10]:


from urllib.parse import urlencode
i=2

#refresh_data dict will have all the city names and their position the sheet that needs to be updated every second
#the format of refresh_data is {city_name:{'row':x,col:'y'}}
refresh_data={}

for k in data:
    
    #all the retriving from url work starts here
    print('Retriving....')
    d={'q':k,'appid':'30a83249bf5fdff79bfd070c89cd7797'} #this is used to make query to the url
    qstr = urlencode(d) #this creates the query
    #next line reads the data from the url
    try:
        result=urllib.request.urlopen('http://api.openweathermap.org/data/2.5/weather?'+qstr).read()
    except:
        sheet.cell(row=i,column=1).value=k
        sheet.cell(row=i,column=2).value='Not Found'
        sheet.cell(row=i,column=3).value='Not Found'
        sheet.cell(row=i,column=4).value=data[k][0]
        sheet.cell(row=i,column=5).value=data[k][1]
        continue
    
    #now converting the data from json to python dict
    jsn=json.loads(result)
    
    celsius_temp=jsn['main']['temp']-273.15 #temp comes in kelvin, so converting to degree celsius
    humidity=jsn['main']['humidity'] #storing the humidity
    
    print('{}: {} {}'.format(k,int(celsius_temp),humidity)) #just to check if everything is going well
    
    sheet.cell(row=i,column=1).value=k #this stores city names to the first column
    
    #storing the temp in the specified format
    if data[k][0]=='C':
        sheet.cell(row=i,column=2).value=int(celsius_temp)
    else:
        sheet.cell(row=i,column=2).value=int((celsius_temp * 9/5) + 32)
    
    sheet.cell(row=i,column=3).value=humidity #storing the humidity
    sheet.cell(row=i,column=4).value=data[k][0]
    sheet.cell(row=i,column=5).value=data[k][1]
    if data[k][1]==1:
        refresh_data[k]={'row':i, 'col':1}
    i=i+1
wb.save('task1.xlsx')


# In[13]:


#everything after this is for refreshing every second until you make it stop

import time

def refresh():
    print('Refreshing...',end=' ')
    for city_name in refresh_data:
        d={'q':city_name,'appid':'30a83249bf5fdff79bfd070c89cd7797'}
        qstr = urlencode(d)
        result=urllib.request.urlopen('http://api.openweathermap.org/data/2.5/weather?'+qstr).read()
        jsn=json.loads(result)
    
        celsius_temp=jsn['main']['temp']-273.15 
        humidity=jsn['main']['humidity']
        
        row_no=refresh_data[city_name]['row']
        
        if sheet.cell(row=row_no, column=4).value=='C': #since temp format is stored in 4th column
            sheet.cell(row=row_no,column=2).value=int(celsius_temp) #since temp is stored in 2nd column
        else:
            sheet.cell(row=row_no,column=2).value=int((celsius_temp * 9/5) + 32)
    wb.save('task1.xlsx')
    
    time.sleep(1)

print("""To stop refreshing, enter 'x'
To continue, enter anything
*This will be asked every 60th time""")
i=0
while True:  
    if i%60==0:
        s = input()
    if s=='x':
        print('Done')
        break
    refresh()
    i+=1
    print(i%60)

